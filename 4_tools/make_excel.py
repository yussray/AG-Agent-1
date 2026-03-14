# ============================================================
# make_excel.py — Antigravity Multi-Agent System
# Generate Excel spreadsheets with openpyxl
# ============================================================

import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import paths  # noqa: E402

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from config import OUTPUT_DIR


def make_excel(
    headers: list,
    rows: list,
    title: str = "Report",
    summary: dict = None,
    insights: list = None,
    filename: str = None
) -> str:
    """
    Generate a formatted Excel file.

    Args:
        headers: Column header names
        rows: List of row value lists
        title: Workbook/sheet title
        summary: Dict of summary metrics for Sheet 2
        insights: List of insight strings for Sheet 3
        filename: Output filename

    Returns:
        Absolute path to saved .xlsx file
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not filename:
        slug = re.sub(r'[^\w]', '_', title.lower())[:25]
        filename = f"excel_{slug}.xlsx"

    path = os.path.join(OUTPUT_DIR, filename)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data ---
    ws = wb.active
    ws.title = "Data"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    if summary:
        ws2 = wb.create_sheet("Summary")
        ws2.cell(1, 1, "Metric").font = Font(bold=True)
        ws2.cell(1, 2, "Value").font = Font(bold=True)
        for i, (k, v) in enumerate(summary.items(), 2):
            ws2.cell(i, 1, k)
            ws2.cell(i, 2, str(v))

    # --- Sheet 3: Insights ---
    if insights:
        ws3 = wb.create_sheet("Insights")
        ws3.cell(1, 1, "Key Insights").font = Font(bold=True, size=12)
        for i, insight in enumerate(insights, 2):
            ws3.cell(i, 1, f"• {insight}")

    wb.save(path)
    return os.path.abspath(path)
